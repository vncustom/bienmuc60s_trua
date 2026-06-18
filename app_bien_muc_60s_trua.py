import datetime
import logging
import os
import re
import threading
import tkinter as tk
import unicodedata
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
from openpyxl.styles import Font
from striprtf.striprtf import rtf_to_text as _rtf_to_text_raw


APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(APP_DIR, "app_bien_muc_60s_trua.log")
APP_TITLE = "Tự động biên mục 60s trưa"
A090_PLACEHOLDER = "K303324"

TRUA_CONFIG = {
    "label": "60s trưa",
    "list_prefix": "BT60STRUA_",
    "output_prefix": "Map_BT60STRUA",
    "a911": "Nguyễn Thị Quyên",
}

OUTPUT_WIDTHS = {
    "A": 10.6640625,
    "B": 48.6640625,
    "C": 77.21875,
    "D": 9.109375,
}

CAMERA_CUE_PHRASES = (
    "CẬN TRÁI",
    "CẬN GIỮA",
    "CẬN PHẢI",
    "TRUNG GIỮA",
    "TOÀN PHẢI",
    "TOÀN TRÁI",
    "TOÀN GIỮA",
)


def prune_old_log_lines(log_path: str):
    if not os.path.exists(log_path):
        return
    today_prefix = datetime.datetime.now().strftime("%Y-%m-%d")
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        kept_lines = []
        keep_current_block = False
        for line in lines:
            date_match = re.match(r"^(\d{4}-\d{2}-\d{2})", line)
            if date_match:
                keep_current_block = date_match.group(1) == today_prefix
            if keep_current_block:
                kept_lines.append(line)
        with open(log_path, "w", encoding="utf-8") as f:
            f.writelines(kept_lines)
    except Exception:
        pass


prune_old_log_lines(LOG_FILE)
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
)


def rtf_to_text(rtf_raw: str) -> str:
    text = _rtf_to_text_raw(rtf_raw)
    return text.replace("\u00f0", "đ").replace("\u00d0", "Đ").replace("\x00", "")


def clean_line(line: str) -> str:
    return re.sub(r"\s+", " ", line.replace("\xa0", " ")).strip()


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def normalize_vietnamese(value: str) -> str:
    value = value.replace("đ", "d").replace("Đ", "D")
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def is_numeric_id(value) -> bool:
    if value is None:
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return bool(re.match(r"^\d+", str(value).strip()))


def id_to_text(value) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_duration(value) -> str:
    if isinstance(value, datetime.datetime):
        value = value.time()
    if isinstance(value, datetime.time):
        total = value.hour * 60 + value.minute
    elif isinstance(value, datetime.timedelta):
        total = int(value.total_seconds())
    elif isinstance(value, (int, float)):
        total = round(float(value) * 24 * 3600)
    elif isinstance(value, str):
        parts = [int(p) for p in re.findall(r"\d+", value)]
        if len(parts) >= 3:
            total = parts[-3] * 3600 + parts[-2] * 60 + parts[-1]
        elif len(parts) == 2:
            total = parts[0] * 60 + parts[1]
        elif len(parts) == 1:
            total = parts[0]
        else:
            total = 0
    else:
        total = 0
    return f"00:{total // 60:02d}:{total % 60:02d}"


def parse_broadcast_date(ws, list_file_name: str) -> datetime.date:
    first_value = ws.cell(row=1, column=1).value
    if isinstance(first_value, str):
        match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", first_value)
        if match:
            day, month, year = match.groups()
            return datetime.date(int(year), int(month), int(day))

    match = re.search(r"(20\d{2})(\d{2})(\d{2})", list_file_name)
    if match:
        year, month, day = match.groups()
        return datetime.date(int(year), int(month), int(day))

    return datetime.date.today()


def should_take_playlist_row(name, video_id, status) -> bool:
    if not isinstance(name, str):
        return False
    normalized = name.strip().lower()
    normalized_key = normalize_vietnamese(normalized)
    if (
        "coming up" in normalized_key
        or "nhung nguoi thuc hien" in normalized_key
        or " end" in normalized_key
    ):
        return False
    if normalized.startswith("60s w "):
        return False
    return (
        (
            normalized.startswith("60s")
            or normalized.startswith("gat60s ")
            or normalized.startswith("60st")
            or normalized.startswith("live -")
        )
        and is_numeric_id(video_id)
    )


def is_vietnamese_upper_title(line: str) -> bool:
    if len(line) < 8:
        return False
    if re.search(r"[a-zàáảãạăằắẳẵặâầấẩẫậèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ]", line):
        return False
    return bool(re.search(r"[A-ZĐÀÁẢÃẠĂẰẮẲẴẶÂẦẤẨẪẬÈÉẺẼẸÊỀẾỂỄỆÌÍỈĨỊÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴ]", line))


def has_vietnamese_signal(line: str) -> bool:
    return bool(re.search(r"[ĐÀÁẢÃẠĂẰẮẲẴẶÂẦẤẨẪẬÈÉẺẼẸÊỀẾỂỄỆÌÍỈĨỊÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴ]", line))


def is_source_slug(line: str) -> bool:
    upper = line.upper()
    if has_vietnamese_signal(upper):
        return False
    return "/" in upper or " - " in upper or upper.count("-") >= 2


def is_stop_line(line: str) -> bool:
    upper = line.upper()
    stop_prefixes = (
        "KHƯƠNG:",
        "KHƯƠNG :",
        "NGỌC:",
        "NGỌC :",
        "TOÀN ",
        "TRUNG ",
        "CẬN ",
        "NGUỒN ",
        "REUTERS",
        "AFP",
        "BIÊN DỊCH",
        "NGÀY ",
        "TÁC GIẢ",
        "THỰC HIỆN",
        "[PB",
        "PHÁT BIỂU",
    )
    return upper.startswith(stop_prefixes) or set(upper) <= {"=", "-", " "}


def get_green_cf_tag(rtf_raw: str) -> str:
    match = re.search(r"\{\\colortbl([^}]+)\}", rtf_raw)
    if not match:
        return r"\cf1"
    colors = match.group(1).split(";")
    for idx, color_def in enumerate(colors):
        if idx == 0:
            continue
        r_match = re.search(r"\\red(\d+)", color_def)
        g_match = re.search(r"\\green(\d+)", color_def)
        b_match = re.search(r"\\blue(\d+)", color_def)
        if not (r_match and g_match and b_match):
            continue
        red, green, blue = int(r_match.group(1)), int(g_match.group(1)), int(b_match.group(1))
        if red == 0 and green == 128 and blue == 0:
            return f"\\cf{idx}"
    return r"\cf1"


def rtf_fragment_to_text(fragment: str, rtf_prefix: str = r"{\rtf1\ansi ") -> str:
    try:
        return clean_line(rtf_to_text(rtf_prefix + fragment + r"\par}"))
    except Exception:
        return ""


def is_bold_green_paragraph(paragraph_raw: str, green_cf_tag: str) -> bool:
    has_green = green_cf_tag in paragraph_raw
    has_bold = bool(re.search(r"\\b(?!0)(?=[\\\s{])", paragraph_raw))
    return has_green and has_bold


def is_title_candidate(line: str) -> bool:
    if len(line) <= 15:
        return False
    upper = line.upper()
    if any(phrase in upper for phrase in CAMERA_CUE_PHRASES):
        return False
    if is_source_slug(line):
        return False
    if is_stop_line(line):
        return False
    return is_vietnamese_upper_title(line)


def extract_title_from_rtf(rtf_raw: str) -> str:
    green_cf_tag = get_green_cf_tag(rtf_raw)
    body_match = re.search(r"(.*?\\pard)(.*)", rtf_raw, flags=re.DOTALL)
    if body_match:
        rtf_prefix = body_match.group(1)
        rtf_body = body_match.group(2)
    else:
        rtf_prefix = r"{\rtf1\ansi "
        rtf_body = rtf_raw
    paragraphs = re.split(r"\\par(?![a-zA-Z])", rtf_body)
    candidates: list[tuple[int, str]] = []

    for idx, paragraph in enumerate(paragraphs):
        if not is_bold_green_paragraph(paragraph, green_cf_tag):
            continue
        line = rtf_fragment_to_text(paragraph, rtf_prefix)
        if is_title_candidate(line):
            candidates.append((idx, line))

    if not candidates:
        plain_lines = [
            line
            for line in (clean_line(x) for x in rtf_to_text(rtf_raw).replace("\r", "").split("\n"))
            if line
        ]
        for line in plain_lines:
            if is_title_candidate(line):
                return line
        return plain_lines[0] if plain_lines else ""

    start_idx, first = candidates[0]
    title_parts = [first]

    if len(title_parts) < 2 and start_idx + 1 < len(paragraphs):
        next_paragraph = paragraphs[start_idx + 1]
        if is_bold_green_paragraph(next_paragraph, green_cf_tag):
            next_line = rtf_fragment_to_text(next_paragraph, rtf_prefix)
            if is_title_candidate(next_line):
                title_parts.append(next_line)

    return " ".join(title_parts)


def read_rtf_raw(path: str) -> str:
    with open(path, "r", encoding="cp1252", errors="ignore") as f:
        return f.read()


def find_matching_rtf(input_dir: str, playlist_name: str) -> str | None:
    target = normalize_name(playlist_name.rstrip("/"))
    candidates = []
    for file_name in os.listdir(input_dir):
        if not file_name.lower().endswith(".rtf"):
            continue
        stem = os.path.splitext(file_name)[0]
        key = normalize_name(stem)
        if key == target:
            return os.path.join(input_dir, file_name)
        if key in target or target in key:
            candidates.append((abs(len(key) - len(target)), file_name))
    if candidates:
        candidates.sort()
        return os.path.join(input_dir, candidates[0][1])
    return None


def extract_ending_rows(input_dir: str, log) -> list[str]:
    ending_path = os.path.join(input_dir, "ENDING.rtf")
    if not os.path.exists(ending_path):
        raise RuntimeError("Không tìm thấy file ENDING.rtf để lấy thông tin $a500.")

    log("Đọc ekip từ ENDING.rtf")
    text = rtf_to_text(read_rtf_raw(ending_path)).replace("\r", "")
    return [line for line in (clean_line(raw_line) for raw_line in text.split("\n")) if line]


def apply_output_style(ws):
    font_13 = Font(name="Times New Roman", size=13)
    font_14 = Font(name="Times New Roman", size=14)
    for cell in ws[1]:
        cell.font = font_13
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).font = font_13
        if ws.cell(row=row_idx, column=2).value:
            ws.cell(row=row_idx, column=2).font = font_13
        if ws.cell(row=row_idx, column=4).value:
            ws.cell(row=row_idx, column=4).font = font_13
    for cell in ws["C"][1:]:
        cell.font = font_14
    for col, width in OUTPUT_WIDTHS.items():
        ws.column_dimensions[col].width = width


def build_map_trua(input_dir: str, output_dir: str, a090: str, a911: str, log) -> str:
    list_prefix = TRUA_CONFIG["list_prefix"]
    list_files = [
        f
        for f in os.listdir(input_dir)
        if f.upper().startswith(list_prefix.upper()) and f.lower().endswith(".xlsx")
    ]
    if not list_files:
        raise RuntimeError(f"Không tìm thấy file {list_prefix}*.xlsx trong thư mục input.")

    list_files.sort()
    list_path = os.path.join(input_dir, list_files[0])
    log(f"Đọc playlist: {list_files[0]}")

    wb_list = openpyxl.load_workbook(list_path, data_only=True)
    ws_list = wb_list.active
    broadcast_date = parse_broadcast_date(ws_list, list_files[0])

    items = []
    for row in ws_list.iter_rows(values_only=True):
        name = row[0] if len(row) > 0 else None
        video_id = row[2] if len(row) > 2 else None
        status = row[3] if len(row) > 3 else None
        duration = row[5] if len(row) > 5 else None
        if should_take_playlist_row(name, video_id, status):
            items.append(
                {
                    "name": name.strip().rstrip("/"),
                    "id": id_to_text(video_id),
                    "duration": format_duration(duration),
                }
            )

    if not items:
        raise RuntimeError(f"Không lọc được tin {TRUA_CONFIG['label']} nào từ playlist.")

    log(f"Đã lọc {len(items)} tin.")

    for item in items:
        rtf_path = find_matching_rtf(input_dir, item["name"])
        if not rtf_path:
            raise RuntimeError(f"Không tìm thấy file RTF cho: {item['name']}")
        item["title"] = extract_title_from_rtf(read_rtf_raw(rtf_path))
        log(f"{item['id']}: {item['title']}")

    a500_rows = extract_ending_rows(input_dir, log)

    os.makedirs(output_dir, exist_ok=True)
    out_name = f"{TRUA_CONFIG['output_prefix']}_{broadcast_date.year}_ Thang{broadcast_date:%m%d}.xlsx"
    out_path = os.path.join(output_dir, out_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["$a090", "$a500", "$a505", "$a911"])

    max_rows = max(len(items), len(a500_rows))
    for idx in range(1, max_rows + 1):
        a500 = a500_rows[idx - 1] if idx <= len(a500_rows) else None
        a505 = None
        if idx <= len(items):
            item = items[idx - 1]
            a505 = f"{idx:02d} - {item['title']}. Thời lượng: {item['duration']}. ID: {item['id']}"
        ws.append([a090, a500, a505, a911 if idx == 1 else None])

    apply_output_style(ws)
    wb.save(out_path)
    return out_path


class BienMuc60sTruaApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("860x610")

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.a090 = tk.StringVar()
        self.a911 = tk.StringVar(value=TRUA_CONFIG["a911"])
        self.a090_entry = None

        self._build_ui()

    def _build_ui(self):
        frm_title = ttk.LabelFrame(self.root, text="Bản tin", padding=10)
        frm_title.pack(fill="x", padx=10, pady=(10, 0))
        ttk.Label(frm_title, text="60s trưa").pack(anchor="w")

        frm_paths = ttk.LabelFrame(self.root, text="Đường dẫn", padding=10)
        frm_paths.pack(fill="x", padx=10, pady=10)

        ttk.Label(frm_paths, text="Thư mục input").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(frm_paths, textvariable=self.input_dir, width=76).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Button(frm_paths, text="Chọn...", command=self.choose_input).grid(row=0, column=2)

        ttk.Label(frm_paths, text="Thư mục output").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(frm_paths, textvariable=self.output_dir, width=76).grid(row=1, column=1, sticky="ew", padx=6)
        ttk.Button(frm_paths, text="Chọn...", command=self.choose_output).grid(row=1, column=2)
        ttk.Label(
            frm_paths,
            text="(Mặc định tạo folder 'output' trong input nếu để trống)",
            font=("Arial", 8, "italic"),
        ).grid(row=2, column=1, sticky="w", padx=6, pady=(0, 4))
        frm_paths.columnconfigure(1, weight=1)

        frm_note = ttk.LabelFrame(self.root, text="Lưu ý định dạng đầu vào để bóc tách đúng", padding=(10, 6))
        frm_note.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Label(
            frm_note,
            justify="left",
            font=("Arial", 9),
            text=(
                f"• File LIST (Excel): Bắt đầu bằng '{TRUA_CONFIG['list_prefix']}' (.xlsx).\n"
                "  Cột A: tên file bắt đầu bằng '60s', 'gat60s ', '60st' hoặc 'live -'.\n"
                "  Cột C: ID phải bắt đầu bằng số; chữ đứng sau số vẫn được bắt.\n"
                "• File RTF tin tức: tên file nên khớp với Cột A trong LIST để app tìm đúng kịch bản.\n"
                "• Cột $a500: lấy toàn bộ nội dung từ ENDING.rtf."
            ),
        ).pack(anchor="w")

        frm_meta = ttk.LabelFrame(self.root, text="Thông tin Map", padding=10)
        frm_meta.pack(fill="x", padx=10, pady=(0, 10))

        ttk.Label(frm_meta, text="Mã bản tin $a090").grid(row=0, column=0, sticky="w", pady=4)
        self.a090_entry = tk.Entry(frm_meta, textvariable=self.a090, width=24)
        self.a090_entry.grid(row=0, column=1, sticky="w", padx=6)
        self._install_placeholder(self.a090_entry, A090_PLACEHOLDER)

        ttk.Label(frm_meta, text="Người biên mục $a911").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(frm_meta, textvariable=self.a911, width=40).grid(row=1, column=1, sticky="w", padx=6)

        frm_actions = ttk.Frame(self.root, padding=(10, 0))
        frm_actions.pack(fill="x")
        self.btn_start = ttk.Button(frm_actions, text="Bắt đầu biên mục", command=self.start)
        self.btn_start.pack(side="right")

        self.log_box = scrolledtext.ScrolledText(self.root, state="disabled", wrap="word", font=("Consolas", 10))
        self.log_box.pack(fill="both", expand=True, padx=10, pady=10)

    def _install_placeholder(self, entry: tk.Entry, text: str):
        entry.insert(0, text)
        entry.configure(foreground="grey")

        def clear_placeholder(_event=None):
            if entry.get() == text and str(entry.cget("foreground")) == "grey":
                entry.delete(0, tk.END)
                entry.configure(foreground="black")

        def restore_placeholder(_event=None):
            if not entry.get().strip():
                entry.insert(0, text)
                entry.configure(foreground="grey")

        entry.bind("<FocusIn>", clear_placeholder)
        entry.bind("<FocusOut>", restore_placeholder)

    def choose_input(self):
        path = filedialog.askdirectory(title="Chọn thư mục input")
        if path:
            self.input_dir.set(path)

    def choose_output(self):
        path = filedialog.askdirectory(title="Chọn thư mục output")
        if path:
            self.output_dir.set(path)

    def log(self, message: str):
        logging.info(message)
        self.root.after(0, self._append_log, message)

    def _append_log(self, message: str):
        self.log_box.configure(state="normal")
        self.log_box.insert(tk.END, f"[{datetime.datetime.now():%H:%M:%S}] {message}\n")
        self.log_box.see(tk.END)
        self.log_box.configure(state="disabled")

    def show_success_dialog(self, out_path: str):
        top = tk.Toplevel(self.root)
        top.title("Thành công")
        top.resizable(False, False)
        top.transient(self.root)
        top.grab_set()

        frm = ttk.Frame(top, padding=16)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Đã tạo file Map thành công:").pack(anchor="w")
        ttk.Label(frm, text=out_path, wraplength=560).pack(anchor="w", pady=(6, 14))

        buttons = ttk.Frame(frm)
        buttons.pack(anchor="e")

        def open_output_folder():
            folder = os.path.dirname(out_path)
            try:
                os.startfile(folder)
            except Exception as exc:
                messagebox.showerror("Lỗi", f"Không mở được thư mục output:\n{exc}", parent=top)

        ttk.Button(buttons, text="Open output folder", command=open_output_folder).pack(side="left", padx=(0, 8))
        ttk.Button(buttons, text="OK", command=top.destroy).pack(side="left")

        top.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - top.winfo_width()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - top.winfo_height()) // 2
        top.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def start(self):
        input_dir = self.input_dir.get().strip()
        if not input_dir or not os.path.isdir(input_dir):
            messagebox.showerror("Lỗi", "Vui lòng chọn thư mục input hợp lệ.")
            return

        output_dir = self.output_dir.get().strip() or os.path.join(input_dir, "output")
        a090 = self.a090.get().strip()
        if self.a090_entry and a090 == A090_PLACEHOLDER and str(self.a090_entry.cget("foreground")) == "grey":
            a090 = ""
        if not a090:
            messagebox.showerror("Lỗi", "Vui lòng nhập mã $a090.")
            return

        self.btn_start.configure(state="disabled")
        self.log(f"Bắt đầu biên mục {TRUA_CONFIG['label']}...")

        def worker():
            try:
                out_path = build_map_trua(input_dir, output_dir, a090, self.a911.get().strip(), self.log)
                self.log(f"Hoàn tất: {out_path}")
                self.root.after(0, lambda: self.show_success_dialog(out_path))
            except Exception as exc:
                logging.exception("Lỗi khi biên mục %s", TRUA_CONFIG["label"])
                self.log(f"LỖI: {exc}")
                self.root.after(0, lambda: messagebox.showerror("Lỗi", str(exc)))
            finally:
                self.root.after(0, lambda: self.btn_start.configure(state="normal"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    app = BienMuc60sTruaApp(root)
    root.mainloop()
